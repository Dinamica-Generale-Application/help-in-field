"""
Aggiunge la colonna "Garanzia" al file Excel esistente.
La colonna viene inserita dopo "price" (colonna W -> X).

Uso:
    python aggiungi_colonna_garanzia.py
"""

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("❌ openpyxl non installato. Esegui: pip install openpyxl")
    input("Premi INVIO per chiudere...")
    sys.exit(1)


def main():
    folder = Path(__file__).parent
    
    # Se eseguito dalla cartella dashboard, usa la cartella PDF
    if folder.name == "dashboard":
        folder = Path(r'H:\DG_Assistenza\Assistenze\2026_Assistenze_in campo')
    
    excel_path = folder / "rapporti.xlsx"
    
    if not excel_path.exists():
        print(f"❌ File non trovato: {excel_path}")
        input("Premi INVIO per chiudere...")
        return
    
    print(f"📂 Apertura: {excel_path}")
    
    # Carica il workbook
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Verifica le intestazioni attuali
    headers = [cell.value for cell in ws[1]]
    print(f"   Colonne attuali: {len(headers)}")
    
    # Trova la posizione di "price"
    try:
        price_col = headers.index("price") + 1  # 1-indexed
    except ValueError:
        print("❌ Colonna 'price' non trovata!")
        input("Premi INVIO per chiudere...")
        return
    
    # Verifica se "Garanzia" esiste già
    if "Garanzia" in headers:
        print("✅ La colonna 'Garanzia' esiste già!")
        input("Premi INVIO per chiudere...")
        return
    
    # Inserisci nuova colonna dopo "price"
    new_col = price_col + 1  # Colonna X (dopo W=price)
    print(f"   Inserimento colonna 'Garanzia' in posizione {new_col} (colonna {chr(64+new_col)})")
    
    ws.insert_cols(new_col)
    
    # Aggiungi intestazione con stile
    header_cell = ws.cell(row=1, column=new_col, value="Garanzia")
    header_cell.font = Font(bold=True, color="FFFFFF")
    header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Imposta larghezza colonna
    ws.column_dimensions[chr(64+new_col)].width = 10
    
    # Aggiungi bordi alle celle dati
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=new_col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Salva
    wb.save(excel_path)
    wb.close()
    
    print(f"✅ Colonna 'Garanzia' aggiunta con successo!")
    print(f"   Ora puoi compilare la colonna con 'Sì' o 'No'")
    
    input("\nPremi INVIO per chiudere...")


if __name__ == "__main__":
    main()
