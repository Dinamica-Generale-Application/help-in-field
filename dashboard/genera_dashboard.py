"""
Dashboard Assistenza — Legge i PDF testuali dei rapporti e genera una dashboard HTML.

Uso:
    Doppio click su questo file, oppure:
    python genera_dashboard.py

Legge tutti i file rapporto_*.pdf nella stessa cartella e genera dashboard.html.
Requisiti: Python 3.10+, PyMuPDF (pip install pymupdf)
"""

import json
import os
import re
import sys
from pathlib import Path
from datetime import datetime

try:
    import fitz  # PyMuPDF
except ImportError:
    print("❌ PyMuPDF non installato. Esegui: pip install pymupdf")
    input("Premi INVIO per chiudere...")
    sys.exit(1)


def load_corrections(folder: Path) -> dict:
    """Carica correzioni manuali da correzioni.json se esiste."""
    corrections_file = folder / "correzioni.json"
    if corrections_file.exists():
        try:
            import json
            with open(corrections_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # Rimuovi chiavi che iniziano con _
                return {k: v for k, v in data.items() if not k.startswith('_')}
        except Exception:
            pass
    return {}


def apply_corrections(data: dict, corrections: dict) -> dict:
    """Applica correzioni manuali ai dati estratti."""
    filename = data.get('_source', '')
    if filename in corrections:
        for key, value in corrections[filename].items():
            data[key] = value
    return data


def normalize_on_behalf(value: str) -> str:
    """Normalizza varianti di 'Fyeld' in un unico valore."""
    if re.match(r"(?i)fyeld", value.strip()):
        return "Fyeld"
    return value.strip()


def normalize_operator(value: str) -> str:
    """Normalizza nomi operatori per unificare varianti (es. 'Bocchino Angelo' -> 'Angelo Bocchino')."""
    if not value:
        return ""
    v = value.strip()
    v_lower = v.lower()
    # Unifica varianti di Angelo Bocchino
    if 'angelo' in v_lower and 'bocchino' in v_lower:
        return "Angelo Bocchino"
    # Unifica varianti di Giacomo Mantovani
    if 'giacomo' in v_lower and 'mantovani' in v_lower:
        return "Giacomo Mantovani"
    # Per altri operatori, normalizza capitalizzazione (Title Case)
    return v.title()


def normalize_company_name(value: str) -> str:
    """Rimuove prefissi comuni come 'Az. Agricola', 'Azienda Agricola', etc."""
    if not value:
        return ""
    v = value.strip()
    # Pattern da rimuovere (case insensitive)
    prefixes = [
        r"^az\.?\s*agricola\s*",
        r"^azienda\s+agricola\s*",
        r"^società\s+agricola\s*",
        r"^soc\.?\s*agricola\s*",
        r"^azienda\s+agricola\s+zootecnica\s*",
        r"^az\.?\s*agr\.?\s*",
    ]
    for pattern in prefixes:
        v = re.sub(pattern, "", v, flags=re.IGNORECASE)
    # Rimuovi spazi multipli e capitalizza
    v = re.sub(r"\s+", " ", v).strip()
    return v.title() if v else value.strip().title()


def extract_report_data(pdf_path: Path) -> dict | None:
    """Estrae i dati strutturati da un PDF testuale generato dalla webapp."""
    try:
        doc = fitz.open(str(pdf_path))
        text = ""
        for page in doc:
            text += page.get_text()
        doc.close()
    except Exception:
        return None

    if not text.strip():
        return None  # PDF immagine (vecchio formato), skip

    data: dict = {"_source": pdf_path.name}

    # Parse fields using label: value pattern
    # In jsPDF PDFs, format can be "Label:\nValue\n" (value on next line)
    # or "Label: Value\n" (value on same line)
    def find_field(label: str) -> str | None:
        # Pattern: "Label:\n" followed by value on next line
        pattern = rf"{re.escape(label)}:\n(.+?)(?:\n|$)"
        match = re.search(pattern, text)
        if not match:
            return None
        value = match.group(1).strip()
        # If value looks like another section/label (contains ":"), it means field was empty
        if not value or value.endswith(":") or value in ("Dati Cliente", "Dettagli Intervento", "Dispositivi", "Costi", "Allegati", "Firma e Timbro"):
            return None
        return value

    data["companyName"] = normalize_company_name(find_field("Ragione Sociale") or "")
    data["address"] = find_field("Indirizzo")
    data["phone"] = find_field("Telefono")

    # Date
    date_str = find_field("Data")  # DD/MM/YYYY or Data Intervento
    if not date_str:
        date_str = find_field("Data Intervento")
    if date_str:
        # Convert DD/MM/YYYY to YYYY-MM-DD for sorting
        match = re.match(r"(\d{2})/(\d{2})/(\d{4})", date_str)
        if match:
            data["interventionDate"] = f"{match.group(3)}-{match.group(2)}-{match.group(1)}"
        else:
            data["interventionDate"] = date_str
    else:
        data["interventionDate"] = ""

    data["operator1"] = normalize_operator(find_field("Operatore 1") or "")
    data["operator2"] = normalize_operator(find_field("Operatore 2") or "")
    data["interventionLocation"] = find_field("Luogo") or find_field("Luogo Intervento")
    data["requestedBy"] = find_field("Richiesto da")
    data["onBehalfOf"] = normalize_on_behalf(find_field("Per conto di") or "Fyeld")
    data["interventionReason"] = find_field("Motivo") or find_field("Motivo richiesta intervento") or find_field("Motivo Intervento")
    data["problemFound"] = find_field("Problema riscontrato")
    data["heatRisk"] = find_field("Rischio Caldo")
    data["description"] = find_field("Descrizione") or find_field("Descrizione dettagliata")
    data["notes"] = find_field("Note")

    # Parse costs from table
    hours_match = re.search(r"Ore lavorate\s*(\d+[.,]?\d*)\s*ore", text)
    if hours_match:
        data["hoursWorked"] = float(hours_match.group(1).replace(",", "."))
    else:
        data["hoursWorked"] = 0

    km_match = re.search(r"Chilometri\s*(\d+[.,]?\d*)\s*km", text)
    if km_match:
        data["kilometers"] = float(km_match.group(1).replace(",", "."))
    else:
        data["kilometers"] = 0

    # Grand total - supporta formati: "156,00 €", "156,00 EUR", "171.78 ·"
    # Pattern 1: valore sulla stessa riga (vecchio formato)
    total_match = re.search(r"Totale Intervento\s*([\d.,]+)\s*(?:€|EUR)", text)
    # Pattern 2: valore su riga successiva (nuovo formato)
    if not total_match:
        total_match = re.search(r"Totale Intervento\n([\d.,]+)\s*(?:€|EUR|·)", text)
    
    if total_match:
        total_str = total_match.group(1).strip()
        # Gestisce formati IT (156,00) e EN (171.78)
        if '.' in total_str and ',' in total_str:
            # Es: "1.234,56" (IT) -> rimuovo punti, virgola -> punto
            if total_str.rfind(',') > total_str.rfind('.'):
                total_str = total_str.replace('.', '').replace(',', '.')
            else:
                total_str = total_str.replace(',', '')
        elif ',' in total_str:
            total_str = total_str.replace(',', '.')
        # Se solo punto, assumo decimale (es: 171.78)
        try:
            data["grandTotal"] = float(total_str)
        except ValueError:
            data["grandTotal"] = 0
    else:
        data["grandTotal"] = 0

    # Detect if travel cost is already included (new format reports)
    data["hasTravelCost"] = "Ore viaggio" in text

    # Only return if we got at least a company name or date
    if data["companyName"] or data["interventionDate"]:
        return data
    return None


def extract_report_from_image(image_path: Path) -> dict | None:
    """Estrae dati da un'immagine JPEG di un report usando OCR (EasyOCR)."""
    try:
        import easyocr
    except ImportError:
        print("  ⚠️  EasyOCR non installato (pip install easyocr). JPEG ignorati.")
        return None

    # Skip small images (likely signature pages, < 30KB)
    if image_path.stat().st_size < 30000:
        return None

    try:
        if not hasattr(extract_report_from_image, '_reader'):
            print("  ⏳ Inizializzazione OCR (prima volta, potrebbe richiedere qualche secondo)...")
            extract_report_from_image._reader = easyocr.Reader(['it', 'en'], gpu=False, verbose=False)
        
        reader = extract_report_from_image._reader
        results = reader.readtext(str(image_path))
        
        # Build text from OCR results (only confident ones)
        lines = [text for (_, text, conf) in results if conf > 0.4]
        text = "\n".join(lines)
    except Exception:
        return None

    if "Rapporto" not in text and "Ragione Sociale" not in text:
        return None  # Not a report image

    data: dict = {"_source": image_path.name}

    # Parse using same logic as PDF but adapted for OCR output
    # OCR gives us lines in reading order, labels and values may be on same or separate lines
    def find_after(label: str) -> str | None:
        for i, line in enumerate(lines):
            if label.lower() in line.lower():
                # Value might be in same line after ":" or on next line
                if ":" in line:
                    parts = line.split(":", 1)
                    if len(parts) > 1 and parts[1].strip():
                        return parts[1].strip()
                # Try next line
                if i + 1 < len(lines):
                    next_line = lines[i + 1]
                    # Skip if next line is another label
                    if not any(lbl in next_line for lbl in ["Dettagli", "Dati Cliente", "Costi", "Voce", "Dispositivi", "Allegati", "Firma"]):
                        return next_line.strip()
        return None

    data["companyName"] = normalize_company_name(find_after("Ragione Sociale") or "")
    
    # Date: look for DD/MM/YYYY pattern
    date_match = re.search(r"(\d{2})/(\d{2})/(\d{4})", text)
    if date_match:
        data["interventionDate"] = f"{date_match.group(3)}-{date_match.group(2)}-{date_match.group(1)}"
    else:
        data["interventionDate"] = ""

    data["operator1"] = find_after("Operatore 1") or ""
    data["operator2"] = find_after("Operatore 2")
    data["interventionLocation"] = find_after("Luogo")
    data["requestedBy"] = find_after("Richiesto da")
    data["onBehalfOf"] = normalize_on_behalf(find_after("Per conto di") or "Fyeld")
    data["interventionReason"] = find_after("Motivo")
    data["problemFound"] = find_after("Problema riscontrato")
    data["description"] = find_after("Descrizione")
    data["notes"] = find_after("Note")

    # Hours: look for "X ore"
    hours_match = re.search(r"(\d+[.,]?\d*)\s*ore", text)
    data["hoursWorked"] = float(hours_match.group(1).replace(",", ".")) if hours_match else 0

    # Km: look for "X km"
    km_match = re.search(r"(\d+[.,]?\d*)\s*km", text)
    data["kilometers"] = float(km_match.group(1).replace(",", ".")) if km_match else 0

    # Total: look for amount after "Totale Intervento" (may be on next line)
    total_match = re.search(r"Totale.*?(\d+[.,]\d+)\s*€", text, re.DOTALL)
    if not total_match:
        # Fallback: find the line after "Totale Intervento"
        for i, line in enumerate(lines):
            if "Totale" in line and "Intervento" in line:
                # Check same line for amount
                amt = re.search(r"(\d+[.,]\d+)\s*€", line)
                if amt:
                    total_match = amt
                # Check next line
                elif i + 1 < len(lines):
                    amt = re.search(r"(\d+[.,]\d+)\s*€", lines[i + 1])
                    if amt:
                        total_match = amt
                break
    if total_match:
        total_str = total_match.group(1).replace(".", "").replace(",", ".")
        try:
            data["grandTotal"] = float(total_str)
        except ValueError:
            data["grandTotal"] = 0
    else:
        data["grandTotal"] = 0

    # JPEG images are always old format (no travel cost included)
    data["hasTravelCost"] = False

    if data["companyName"] or data["interventionDate"]:
        return data
    return None


def load_from_excel(excel_path: Path) -> list[dict]:
    """Carica i dati dal file Excel compilato."""
    try:
        import openpyxl
    except ImportError:
        print("❌ openpyxl non installato. Esegui: pip install openpyxl")
        return []
    
    if not excel_path.exists():
        return []
    
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    
    reports = []
    headers = [cell.value for cell in ws[1]]
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Skip empty rows
            continue
        
        data = {}
        data["_source"] = row[0] or ""
        data["interventionDateDisplay"] = row[1] or ""
        
        # Converti data display in formato ISO per sorting
        date_str = row[1] or ""
        if date_str:
            match = re.match(r"(\d{2})/(\d{2})/(\d{4})", str(date_str))
            if match:
                data["interventionDate"] = f"{match.group(3)}-{match.group(2)}-{match.group(1)}"
            else:
                data["interventionDate"] = str(date_str)
        else:
            data["interventionDate"] = ""
        
        data["companyName"] = row[2] or ""
        data["address"] = row[3] or ""
        data["phone"] = row[4] or ""
        data["operator1"] = row[5] or ""
        data["operator2"] = row[6] or ""
        data["onBehalfOf"] = row[7] or "Dinamica Generale"
        data["interventionReason"] = row[8] or ""
        data["problemFound"] = row[9] or ""
        data["description"] = row[10] or ""
        data["deviceType"] = row[11] or ""
        data["serial1"] = row[12] or ""
        data["year1"] = row[13] or ""
        data["serial2"] = row[14] or ""
        data["year2"] = row[15] or ""
        data["serial3"] = row[16] or ""
        data["year3"] = row[17] or ""
        data["serial4"] = row[18] or ""
        data["year4"] = row[19] or ""
        data["spareParts"] = row[20] or ""
        data["sparePartsQty"] = row[21] or ""
        data["sparePartsPrice"] = float(row[22] or 0) if row[22] else 0
        data["warranty"] = str(row[23] or "").strip()
        data["hoursWorked"] = float(row[24] or 0)
        data["kilometers"] = float(row[25] or 0)
        data["grandTotal"] = float(row[26] or 0)
        data["hasTravelCost"] = str(row[27] or "").lower() == "sì"
        data["notes"] = row[28] or ""
        data["notes"] = row[25] or ""
        
        reports.append(data)
    
    wb.close()
    return reports


def load_reports(folder: Path) -> list[dict]:
    """Carica dati da Excel se esiste, altrimenti dai PDF."""
    
    # Prima prova a caricare dall'Excel
    excel_path = folder / "rapporti.xlsx"
    if excel_path.exists():
        print(f"  📊 Caricamento da Excel: {excel_path.name}")
        reports = load_from_excel(excel_path)
        if reports:
            # Verifica se ci sono nuovi PDF non presenti nell'Excel
            excel_files = set(r.get("_source", "") for r in reports)
            pdf_files = set(p.name for p in folder.glob("rapporto_*.pdf"))
            new_pdfs = pdf_files - excel_files
            
            if new_pdfs:
                print(f"  ⚠️  ATTENZIONE: {len(new_pdfs)} nuovi PDF non presenti nell'Excel:")
                for pdf in sorted(new_pdfs)[:10]:  # Mostra max 10
                    print(f"      - {pdf}")
                if len(new_pdfs) > 10:
                    print(f"      ... e altri {len(new_pdfs) - 10}")
                print(f"  💡 Esegui genera_excel.py per aggiornare il file Excel")
            
            print(f"  ✅ Caricati {len(reports)} rapporti da Excel")
            return reports
        print(f"  ⚠️  Excel vuoto, fallback a PDF")
    
    # Fallback: carica dai PDF
    reports = []
    
    # Carica correzioni manuali
    corrections = load_corrections(folder)
    if corrections:
        print(f"  📝 Caricate {len(corrections)} correzioni da correzioni.json")
    
    # Load from PDFs
    pdf_files = sorted(folder.glob("rapporto_*.pdf"))
    for pdf_path in pdf_files:
        data = extract_report_data(pdf_path)
        if data:
            # Applica correzioni manuali
            data = apply_corrections(data, corrections)
            reports.append(data)
        else:
            print(f"  ⚠️  Ignorato (vecchio formato o vuoto): {pdf_path.name}")

    return reports


def generate_dashboard_html(reports: list[dict]) -> str:
    """Genera l'HTML della dashboard con dati embedded come JSON."""
    reports_json = json.dumps(reports, ensure_ascii=False, indent=None)
    generated_at = datetime.now().strftime("%d/%m/%Y %H:%M")

    return f"""<!DOCTYPE html>
<html lang="it">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Dashboard Assistenza — Dinamica Generale</title>
<style>
  * {{ margin: 0; padding: 0; box-sizing: border-box; }}
  body {{
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    background: #f5f7fa;
    color: #1a1a2e;
    padding: 20px;
    line-height: 1.5;
  }}
  h1 {{
    text-align: center;
    font-size: 1.8rem;
    margin-bottom: 5px;
    color: #1a1a2e;
  }}
  .subtitle {{
    text-align: center;
    color: #666;
    margin-bottom: 25px;
    font-size: 0.85rem;
  }}
  .filters {{
    background: white;
    border-radius: 12px;
    padding: 20px;
    margin-bottom: 20px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    display: flex;
    flex-wrap: wrap;
    gap: 15px;
    align-items: flex-end;
  }}
  .filter-group {{
    display: flex;
    flex-direction: column;
    gap: 4px;
  }}
  .filter-group label {{
    font-size: 0.75rem;
    font-weight: 600;
    color: #555;
    text-transform: uppercase;
    letter-spacing: 0.5px;
  }}
  .filter-group input, .filter-group select {{
    padding: 8px 12px;
    border: 1px solid #ddd;
    border-radius: 6px;
    font-size: 0.9rem;
    min-width: 150px;
  }}
  .filter-group input:focus, .filter-group select:focus {{
    outline: none;
    border-color: #4a90d9;
    box-shadow: 0 0 0 2px rgba(74,144,217,0.2);
  }}
  .btn-reset {{
    padding: 8px 16px;
    background: #e8e8e8;
    border: none;
    border-radius: 6px;
    cursor: pointer;
    font-size: 0.85rem;
    font-weight: 500;
  }}
  .btn-reset:hover {{ background: #ddd; }}
  .kpi-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    gap: 15px;
    margin-bottom: 25px;
  }}
  .kpi-card {{
    background: white;
    border-radius: 12px;
    padding: 20px;
    text-align: center;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
  }}
  .kpi-value {{
    font-size: 2rem;
    font-weight: 700;
    color: #2563eb;
  }}
  .kpi-label {{
    font-size: 0.8rem;
    color: #666;
    margin-top: 4px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
  }}
  /* Tab styling for expense split */
  .expense-tabs {{
    display: flex;
    gap: 0;
    margin-bottom: 8px;
    justify-content: center;
  }}
  .expense-tab {{
    padding: 6px 12px;
    font-size: 0.7rem;
    font-weight: 600;
    cursor: pointer;
    border: 1px solid #ddd;
    background: #f5f5f5;
    color: #666;
    text-transform: uppercase;
    letter-spacing: 0.3px;
    transition: all 0.2s;
  }}
  .expense-tab:first-child {{
    border-radius: 6px 0 0 6px;
  }}
  .expense-tab:last-child {{
    border-radius: 0 6px 6px 0;
    border-left: none;
  }}
  .expense-tab.active {{
    background: #2563eb;
    color: white;
    border-color: #2563eb;
  }}
  .expense-tab:hover:not(.active) {{
    background: #e8e8e8;
  }}
  .charts-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(400px, 1fr));
    gap: 20px;
    margin-bottom: 25px;
  }}
  .chart-card {{
    background: white;
    border-radius: 12px;
    padding: 20px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
  }}
  .chart-card h3 {{
    font-size: 0.9rem;
    color: #555;
    margin-bottom: 15px;
    text-transform: uppercase;
    letter-spacing: 0.5px;
  }}
  table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 0.85rem;
  }}
  th {{
    text-align: left;
    padding: 10px 12px;
    background: #f8f9fa;
    font-weight: 600;
    color: #555;
    border-bottom: 2px solid #e8e8e8;
  }}
  td {{
    padding: 10px 12px;
    border-bottom: 1px solid #f0f0f0;
  }}
  tr:hover td {{ background: #f8f9ff; }}
  .bar-container {{
    display: flex;
    align-items: center;
    gap: 8px;
    margin-bottom: 8px;
  }}
  .bar-label {{
    min-width: 120px;
    font-size: 0.8rem;
    text-align: right;
    color: #555;
  }}
  .bar {{
    height: 24px;
    background: linear-gradient(90deg, #4a90d9, #64b5f6);
    border-radius: 4px;
    transition: width 0.3s ease;
    min-width: 2px;
  }}
  .bar-value {{
    font-size: 0.8rem;
    font-weight: 600;
    color: #333;
    min-width: 30px;
  }}
  .no-data {{
    text-align: center;
    color: #999;
    padding: 40px;
    font-style: italic;
  }}
  .bar-clickable {{
    cursor: pointer;
    transition: opacity 0.2s;
  }}
  .bar-clickable:hover {{
    opacity: 0.8;
  }}
  /* Modal */
  .modal-overlay {{
    display: none;
    position: fixed;
    top: 0;
    left: 0;
    width: 100%;
    height: 100%;
    background: rgba(0,0,0,0.5);
    z-index: 1000;
    justify-content: center;
    align-items: center;
  }}
  .modal-overlay.active {{
    display: flex;
  }}
  .modal-content {{
    background: white;
    border-radius: 12px;
    padding: 25px;
    max-width: 90%;
    max-height: 80%;
    overflow-y: auto;
    box-shadow: 0 10px 40px rgba(0,0,0,0.2);
    min-width: 600px;
  }}
  .modal-header {{
    display: flex;
    justify-content: space-between;
    align-items: center;
    margin-bottom: 20px;
    padding-bottom: 15px;
    border-bottom: 1px solid #eee;
  }}
  .modal-header h2 {{
    font-size: 1.2rem;
    color: #333;
  }}
  .modal-close {{
    background: none;
    border: none;
    font-size: 1.5rem;
    cursor: pointer;
    color: #999;
    padding: 5px;
  }}
  .modal-close:hover {{
    color: #333;
  }}
  .modal-table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 0.85rem;
  }}
  .modal-table th {{
    text-align: left;
    padding: 10px;
    background: #f8f9fa;
    font-weight: 600;
    border-bottom: 2px solid #e8e8e8;
  }}
  .modal-table td {{
    padding: 10px;
    border-bottom: 1px solid #f0f0f0;
  }}
  .modal-table tr:hover td {{
    background: #f8f9ff;
  }}
  @media (max-width: 768px) {{
    .charts-grid {{ grid-template-columns: 1fr; }}
    .filters {{ flex-direction: column; }}
    .modal-content {{ min-width: auto; width: 95%; }}
  }}
</style>
</head>
<body>

<!-- Modal per dettaglio interventi -->
<div class="modal-overlay" id="detailModal" onclick="closeModal(event)">
  <div class="modal-content" onclick="event.stopPropagation()">
    <div class="modal-header">
      <h2 id="modalTitle">Dettaglio Interventi</h2>
      <button class="modal-close" onclick="closeModal()">&times;</button>
    </div>
    <div id="modalBody"></div>
  </div>
</div>

<h1>📊 Dashboard Assistenza</h1>
<p class="subtitle">Generata il {generated_at} — {len(reports)} rapporti caricati</p>

<!-- Filtri -->
<div class="filters">
  <div class="filter-group">
    <label>Da</label>
    <input type="date" id="filterFrom" />
  </div>
  <div class="filter-group">
    <label>A</label>
    <input type="date" id="filterTo" />
  </div>
  <div class="filter-group">
    <label>Cliente</label>
    <select id="filterCompany"><option value="">Tutti</option></select>
  </div>
  <div class="filter-group">
    <label>Richiesto da</label>
    <select id="filterRequestedBy"><option value="">Tutti</option></select>
  </div>
  <div class="filter-group">
    <label>Per conto di</label>
    <select id="filterOnBehalfOf"><option value="">Tutti</option></select>
  </div>
  <div class="filter-group">
    <label>Motivo</label>
    <select id="filterReason">
      <option value="">Tutti</option>
      <option value="Installazione">Installazione</option>
      <option value="Supervisione">Supervisione</option>
      <option value="Malfunzionamento">Malfunzionamento</option>
      <option value="Altro">Altro</option>
    </select>
  </div>
  <div class="filter-group">
    <label>Operatore</label>
    <select id="filterOperator"><option value="">Tutti</option></select>
  </div>
  <button class="btn-reset" onclick="resetFilters()">↺ Reset</button>
</div>

<!-- KPI Cards -->
<div class="kpi-grid" id="kpiGrid"></div>

<!-- Charts -->
<div class="charts-grid">
  <div class="chart-card">
    <h3>Interventi per mese</h3>
    <div id="chartMonthly"></div>
  </div>
  <div class="chart-card">
    <h3>Top clienti</h3>
    <div id="chartClients"></div>
  </div>
  <div class="chart-card">
    <h3>Per problema riscontrato</h3>
    <div id="chartProblems"></div>
  </div>
  <div class="chart-card">
    <h3>Interventi in garanzia</h3>
    <div id="chartWarranty"></div>
  </div>
  <div class="chart-card">
    <h3>Dispositivi per anno di produzione</h3>
    <div id="chartDeviceYears"></div>
  </div>
  <div class="chart-card">
    <h3>Problema riscontrato per anno dispositivo</h3>
    <div id="chartReasonByYear"></div>
  </div>
</div>

<script>
const ALL_REPORTS = {reports_json};

// --- Populate filter dropdowns ---
function populateFilters() {{
  const companies = [...new Set(ALL_REPORTS.map(r => r.companyName).filter(Boolean))].sort();
  const requestedBys = [...new Set(ALL_REPORTS.map(r => r.requestedBy).filter(Boolean))].sort();
  const onBehalfOfs = [...new Set(ALL_REPORTS.map(r => r.onBehalfOf).filter(Boolean))].sort();
  const operators = [...new Set(ALL_REPORTS.flatMap(r => [r.operator1, r.operator2]).filter(Boolean))].sort();

  fillSelect('filterCompany', companies);
  fillSelect('filterRequestedBy', requestedBys);
  fillSelect('filterOnBehalfOf', onBehalfOfs);
  fillSelect('filterOperator', operators);
}}

function fillSelect(id, values) {{
  const sel = document.getElementById(id);
  const current = sel.value;
  while (sel.options.length > 1) sel.remove(1);
  values.forEach(v => {{
    const opt = document.createElement('option');
    opt.value = v;
    opt.textContent = v;
    sel.appendChild(opt);
  }});
  sel.value = current;
}}

// --- Filter logic ---
function getFilteredReports() {{
  const from = document.getElementById('filterFrom').value;
  const to = document.getElementById('filterTo').value;
  const company = document.getElementById('filterCompany').value;
  const requestedBy = document.getElementById('filterRequestedBy').value;
  const onBehalfOf = document.getElementById('filterOnBehalfOf').value;
  const reason = document.getElementById('filterReason').value;
  const operator = document.getElementById('filterOperator').value;

  return ALL_REPORTS.filter(r => {{
    if (from && r.interventionDate < from) return false;
    if (to && r.interventionDate > to) return false;
    if (company && r.companyName !== company) return false;
    if (requestedBy && r.requestedBy !== requestedBy) return false;
    if (onBehalfOf && r.onBehalfOf !== onBehalfOf) return false;
    if (reason && r.interventionReason !== reason) return false;
    if (operator && r.operator1 !== operator && r.operator2 !== operator) return false;
    return true;
  }});
}}

function resetFilters() {{
  document.getElementById('filterFrom').value = '';
  document.getElementById('filterTo').value = '';
  document.getElementById('filterCompany').value = '';
  document.getElementById('filterRequestedBy').value = '';
  document.getElementById('filterOnBehalfOf').value = '';
  document.getElementById('filterReason').value = '';
  document.getElementById('filterOperator').value = '';
  render();
}}

// --- Render ---
function render() {{
  const reports = getFilteredReports();
  renderKPIs(reports);
  renderMonthlyChart(reports);
  renderClientsChart(reports);
  renderProblemsChart(reports);
  renderWarrantyChart(reports);
  renderDeviceYearsChart(reports);
  renderReasonByYearChart(reports);
}}

// Helper: check if report belongs to Angelo Bocchino
// I nomi operatori sono già normalizzati in Python, quindi basta confrontare
function isAngeloBocchino(report) {{
  const op1 = (report.operator1 || '').trim();
  const op2 = (report.operator2 || '').trim();
  return op1 === 'Angelo Bocchino' || op2 === 'Angelo Bocchino';
}}

// Calculate total expenses for a set of reports (include spare parts cost)
function calcExpenses(reports) {{
  let total = 0;
  reports.forEach(r => {{
    let reportTotal = r.grandTotal || 0;
    if (!r.hasTravelCost && r.kilometers > 0) {{
      reportTotal += (r.kilometers / 55) * 60;
    }}
    // Aggiungi costo ricambi
    reportTotal += parseFloat(r.sparePartsPrice) || 0;
    total += reportTotal;
  }});
  return total;
}}

// Track current expense tab
let expenseTab = 'angelo'; // 'angelo' or 'altri'

function setExpenseTab(tab) {{
  expenseTab = tab;
  // Re-render just the expense card
  const reports = getFilteredReports();
  updateExpenseCard(reports);
}}

// Compenso fisso Angelo Bocchino per 3 mesi di supporto in campo
const ANGELO_FIXED_COMPENSATION = 20000;

function updateExpenseCard(reports) {{
  const angeloReports = reports.filter(isAngeloBocchino);
  const altriReports = reports.filter(r => !isAngeloBocchino(r));
  
  const angeloExpenses = calcExpenses(angeloReports) + ANGELO_FIXED_COMPENSATION;
  const altriExpenses = calcExpenses(altriReports);
  
  const currentExpenses = expenseTab === 'angelo' ? angeloExpenses : altriExpenses;
  const currentCount = expenseTab === 'angelo' ? angeloReports.length : altriReports.length;
  
  document.getElementById('expenseCard').innerHTML = `
    <div class="expense-tabs">
      <button class="expense-tab ${{expenseTab === 'angelo' ? 'active' : ''}}" onclick="setExpenseTab('angelo')">A. Bocchino (${{angeloReports.length}})</button>
      <button class="expense-tab ${{expenseTab === 'altri' ? 'active' : ''}}" onclick="setExpenseTab('altri')">Altri (${{altriReports.length}})</button>
    </div>
    <div class="kpi-value">${{currentExpenses.toLocaleString('it-IT', {{minimumFractionDigits:2, maximumFractionDigits:2}})}}&euro;</div>
    <div class="kpi-label">Spese totali</div>
    ${{expenseTab === 'angelo' ? '<div style="font-size: 0.7rem; color: #666; margin-top: 4px;">(incl. 20.000€ compenso fisso)</div>' : ''}}
  `;
}}

function renderKPIs(reports) {{
  const totalInterventions = reports.length;
  const totalHours = reports.reduce((s, r) => s + (r.hoursWorked || 0), 0);
  const totalKm = reports.reduce((s, r) => s + (r.kilometers || 0), 0);
  const avgHoursPerIntervention = totalInterventions > 0 ? (totalHours / totalInterventions).toFixed(1) : '0';
  const uniqueClients = new Set(reports.map(r => r.companyName)).size;
  
  // Totale costi ricambi
  const totalSpareParts = reports.reduce((s, r) => s + (parseFloat(r.sparePartsPrice) || 0), 0);
  
  // Helper per verificare garanzia
  const isWarranty = (r) => {{
    const w = (r.warranty || '').toLowerCase().trim();
    return w === 'sì' || w === 'si' || w === 'yes' || w === 's';
  }};
  
  // Costi ricambi per interventi in garanzia
  const warrantySparePartsCost = reports.filter(isWarranty).reduce((s, r) => s + (parseFloat(r.sparePartsPrice) || 0), 0);
  
  // Interventi in garanzia e non
  const warrantyReports = reports.filter(isWarranty);
  const nonWarrantyReports = reports.filter(r => !isWarranty(r));
  
  // Spese totali in garanzia (interventi + ricambi)
  const warrantyExpenses = calcExpenses(warrantyReports);
  
  // Spese totali NON in garanzia (interventi + ricambi)
  const nonWarrantyExpenses = calcExpenses(nonWarrantyReports);

  // Stima ore di viaggio totali: km / 55 km/h
  const travelHours = totalKm / 55;

  document.getElementById('kpiGrid').innerHTML = `
    <div class="kpi-card"><div class="kpi-value">${{totalInterventions}}</div><div class="kpi-label">Interventi</div></div>
    <div class="kpi-card"><div class="kpi-value">${{totalHours.toFixed(1)}}</div><div class="kpi-label">Ore lavorate</div></div>
    <div class="kpi-card"><div class="kpi-value">${{travelHours.toFixed(1)}}</div><div class="kpi-label">Ore viaggio (stima)</div></div>
    <div class="kpi-card"><div class="kpi-value">${{totalKm.toLocaleString('it-IT')}}</div><div class="kpi-label">Km totali</div></div>
    <div class="kpi-card">
      <div class="kpi-value">${{totalSpareParts.toLocaleString('it-IT', {{minimumFractionDigits: 2, maximumFractionDigits: 2}})}}&euro;</div>
      <div class="kpi-label">Costo ricambi</div>
      <div style="font-size: 0.75rem; color: #4CAF50; margin-top: 4px;">di cui in garanzia: ${{warrantySparePartsCost.toLocaleString('it-IT', {{minimumFractionDigits: 2, maximumFractionDigits: 2}})}}&euro;</div>
    </div>
    <div class="kpi-card" id="expenseCard"></div>
    <div class="kpi-card" style="background: linear-gradient(135deg, #e8f5e9 0%, #c8e6c9 100%);">
      <div class="kpi-value" style="color: #2e7d32;">${{warrantyExpenses.toLocaleString('it-IT', {{minimumFractionDigits: 2, maximumFractionDigits: 2}})}}&euro;</div>
      <div class="kpi-label">Spese IN garanzia</div>
      <div style="font-size: 0.7rem; color: #666; margin-top: 4px;">${{warrantyReports.length}} interventi</div>
    </div>
    <div class="kpi-card" style="background: linear-gradient(135deg, #ffebee 0%, #ffcdd2 100%);">
      <div class="kpi-value" style="color: #c62828;">${{nonWarrantyExpenses.toLocaleString('it-IT', {{minimumFractionDigits: 2, maximumFractionDigits: 2}})}}&euro;</div>
      <div class="kpi-label">Spese NON in garanzia</div>
      <div style="font-size: 0.7rem; color: #666; margin-top: 4px;">${{nonWarrantyReports.length}} interventi</div>
    </div>
    <div class="kpi-card"><div class="kpi-value">${{avgHoursPerIntervention}}</div><div class="kpi-label">Ore medie / intervento</div></div>
    <div class="kpi-card"><div class="kpi-value">${{uniqueClients}}</div><div class="kpi-label">Clienti unici</div></div>
  `;
  
  // Render the expense card with tabs
  updateExpenseCard(reports);
}}

function renderBarChart(containerId, data, maxBars, filterType, preserveOrder) {{
  const container = document.getElementById(containerId);
  if (data.length === 0) {{
    container.innerHTML = '<div class="no-data">Nessun dato</div>';
    return;
  }}
  // Se preserveOrder è true, mantieni l'ordine originale, altrimenti ordina per valore
  const sorted = preserveOrder ? data.slice(0, maxBars || 10) : data.sort((a, b) => b.value - a.value).slice(0, maxBars || 10);
  const maxVal = Math.max(...sorted.map(d => d.value), 1);
  container.innerHTML = sorted.map(d => `
    <div class="bar-container bar-clickable" onclick="showDetail('${{filterType || containerId}}', '${{d.label.replace(/'/g, "\\\\'")}}')">
      <div class="bar-label">${{d.label}}</div>
      <div class="bar" style="width: ${{(d.value / maxVal * 100).toFixed(1)}}%"></div>
      <div class="bar-value">${{d.value}}</div>
    </div>
  `).join('');
}}

// Modal functions
function showDetail(filterType, value) {{
  const reports = getFilteredReports();
  let filtered = [];
  let title = '';
  
  switch(filterType) {{
    case 'chartMonthly':
      // value è tipo "Lug 2026" - devo convertire in YYYY-MM
      const monthMap = {{'Gen':'01','Feb':'02','Mar':'03','Apr':'04','Mag':'05','Giu':'06','Lug':'07','Ago':'08','Set':'09','Ott':'10','Nov':'11','Dic':'12'}};
      const parts = value.split(' ');
      if (parts.length === 2) {{
        const mm = monthMap[parts[0]] || '01';
        const yyyy = parts[1];
        const ym = `${{yyyy}}-${{mm}}`;
        filtered = reports.filter(r => r.interventionDate?.startsWith(ym));
      }}
      title = `Interventi di ${{value}}`;
      break;
    case 'chartClients':
      filtered = reports.filter(r => (r.companyName || '').substring(0, 25) === value || r.companyName === value);
      title = `Interventi per ${{value}}`;
      break;
    case 'chartReasons':
      filtered = reports.filter(r => (r.interventionReason || 'Non specificato') === value);
      title = `Interventi: ${{value}}`;
      break;
    case 'chartProblems':
      filtered = reports.filter(r => (r.problemFound || 'Non specificato') === value);
      title = `Problema: ${{value}}`;
      break;
    case 'chartDeviceYears':
      filtered = reports.filter(r => 
        r.year1 == value || r.year2 == value || r.year3 == value || r.year4 == value
      );
      title = `Dispositivi anno ${{value}}`;
      break;
    default:
      filtered = reports;
      title = 'Dettaglio interventi';
  }}
  
  document.getElementById('modalTitle').textContent = `${{title}} (${{filtered.length}})`;
  
  let html = `<table class="modal-table">
    <thead>
      <tr>
        <th>Data</th>
        <th>Cliente</th>
        <th>Motivo</th>
        <th>Problema</th>
        <th>Seriali</th>
        <th>Ore</th>
        <th>Km</th>
        <th>Totale</th>
      </tr>
    </thead>
    <tbody>`;
  
  filtered.sort((a, b) => (b.interventionDate || '').localeCompare(a.interventionDate || ''));
  
  filtered.forEach(r => {{
    const serials = [r.serial1, r.serial2, r.serial3, r.serial4].filter(s => s).join(', ');
    html += `<tr>
      <td>${{formatDate(r.interventionDate)}}</td>
      <td>${{r.companyName || ''}}</td>
      <td>${{r.interventionReason || ''}}</td>
      <td>${{r.problemFound || ''}}</td>
      <td style="font-size: 0.75rem;">${{serials}}</td>
      <td>${{r.hoursWorked || 0}}</td>
      <td>${{r.kilometers || 0}}</td>
      <td>${{(r.grandTotal || 0).toLocaleString('it-IT', {{minimumFractionDigits:2}})}}&euro;</td>
    </tr>`;
  }});
  
  html += '</tbody></table>';
  document.getElementById('modalBody').innerHTML = html;
  document.getElementById('detailModal').classList.add('active');
}}

function showModalWithReports(title, filtered) {{
  document.getElementById('modalTitle').textContent = `${{title}} (${{filtered.length}})`;
  
  let html = `<table class="modal-table">
    <thead>
      <tr>
        <th>Data</th>
        <th>Cliente</th>
        <th>Motivo</th>
        <th>Problema</th>
        <th>Seriali</th>
        <th>Ore</th>
        <th>Km</th>
        <th>Totale</th>
      </tr>
    </thead>
    <tbody>`;
  
  filtered.sort((a, b) => (b.interventionDate || '').localeCompare(a.interventionDate || ''));
  
  filtered.forEach(r => {{
    const serials = [r.serial1, r.serial2, r.serial3, r.serial4].filter(s => s).join(', ');
    html += `<tr>
      <td>${{formatDate(r.interventionDate)}}</td>
      <td>${{r.companyName || ''}}</td>
      <td>${{r.interventionReason || ''}}</td>
      <td>${{r.problemFound || ''}}</td>
      <td style="font-size: 0.75rem;">${{serials}}</td>
      <td>${{r.hoursWorked || 0}}</td>
      <td>${{r.kilometers || 0}}</td>
      <td>${{(r.grandTotal || 0).toLocaleString('it-IT', {{minimumFractionDigits:2}})}}&euro;</td>
    </tr>`;
  }});
  
  html += '</tbody></table>';
  document.getElementById('modalBody').innerHTML = html;
  document.getElementById('detailModal').classList.add('active');
}}

function closeModal(event) {{
  if (!event || event.target.id === 'detailModal') {{
    document.getElementById('detailModal').classList.remove('active');
  }}
}}

// Close modal on Escape key
document.addEventListener('keydown', (e) => {{
  if (e.key === 'Escape') closeModal();
}});

function renderMonthlyChart(reports) {{
  const monthly = {{}};
  reports.forEach(r => {{
    const month = r.interventionDate?.substring(0, 7);
    if (month) monthly[month] = (monthly[month] || 0) + 1;
  }});
  const data = Object.entries(monthly)
    .sort(([a], [b]) => a.localeCompare(b))
    .map(([k, v]) => ({{ label: formatMonth(k), value: v }}));
  renderBarChart('chartMonthly', data, 12, 'chartMonthly');
}}

function renderClientsChart(reports) {{
  const counts = {{}};
  reports.forEach(r => {{
    const name = r.companyName || 'N/D';
    counts[name] = (counts[name] || 0) + 1;
  }});
  const data = Object.entries(counts).map(([k, v]) => ({{ label: k.substring(0, 25), value: v }}));
  renderBarChart('chartClients', data, 8, 'chartClients');
}}

function renderReasonsChart(reports) {{
  const counts = {{}};
  reports.forEach(r => {{
    const reason = r.interventionReason || 'Non specificato';
    counts[reason] = (counts[reason] || 0) + 1;
  }});
  const data = Object.entries(counts).map(([k, v]) => ({{ label: k, value: v }}));
  renderBarChart('chartReasons', data, 5, 'chartReasons');
}}

function renderProblemsChart(reports) {{
  const counts = {{}};
  reports.forEach(r => {{
    const problem = r.problemFound || 'Non specificato';
    counts[problem] = (counts[problem] || 0) + 1;
  }});
  const data = Object.entries(counts).map(([k, v]) => ({{ label: k, value: v }}));
  renderBarChart('chartProblems', data, 10, 'chartProblems');
}}

function renderWarrantyChart(reports) {{
  const container = document.getElementById('chartWarranty');
  
  let inGaranzia = 0;
  let fuoriGaranzia = 0;
  let nonSpecificato = 0;
  
  reports.forEach(r => {{
    const warranty = (r.warranty || '').toLowerCase().trim();
    if (warranty === 'sì' || warranty === 'si' || warranty === 'yes' || warranty === 's') {{
      inGaranzia++;
    }} else if (warranty === 'no' || warranty === 'n') {{
      fuoriGaranzia++;
    }} else {{
      nonSpecificato++;
    }}
  }});
  
  const total = reports.length;
  if (total === 0) {{
    container.innerHTML = '<div class="no-data">Nessun dato</div>';
    return;
  }}
  
  // Calcola percentuali
  const pctGaranzia = (inGaranzia / total * 100).toFixed(1);
  const pctFuori = (fuoriGaranzia / total * 100).toFixed(1);
  const pctNonSpec = (nonSpecificato / total * 100).toFixed(1);
  
  let html = '<div style="display: flex; flex-direction: column; gap: 12px;">';
  
  // Barra In garanzia (verde)
  html += `<div class="bar-container bar-clickable" onclick="showWarrantyDetail('sì')">
    <div class="bar-label">In garanzia</div>
    <div class="bar" style="width: ${{pctGaranzia}}%; background: linear-gradient(90deg, #4CAF50, #81C784);"></div>
    <div class="bar-value">${{inGaranzia}} (${{pctGaranzia}}%)</div>
  </div>`;
  
  // Barra Fuori garanzia (rosso)
  html += `<div class="bar-container bar-clickable" onclick="showWarrantyDetail('no')">
    <div class="bar-label">Fuori garanzia</div>
    <div class="bar" style="width: ${{pctFuori}}%; background: linear-gradient(90deg, #f44336, #e57373);"></div>
    <div class="bar-value">${{fuoriGaranzia}} (${{pctFuori}}%)</div>
  </div>`;
  
  // Barra Non specificato (grigio) - solo se ci sono
  if (nonSpecificato > 0) {{
    html += `<div class="bar-container bar-clickable" onclick="showWarrantyDetail('')">
      <div class="bar-label">Non specificato</div>
      <div class="bar" style="width: ${{pctNonSpec}}%; background: linear-gradient(90deg, #9E9E9E, #BDBDBD);"></div>
      <div class="bar-value">${{nonSpecificato}} (${{pctNonSpec}}%)</div>
    </div>`;
  }}
  
  html += '</div>';
  container.innerHTML = html;
}}

function showWarrantyDetail(warrantyValue) {{
  const reports = getFilteredReports();
  let filtered = [];
  let title = '';
  
  if (warrantyValue === 'sì') {{
    filtered = reports.filter(r => {{
      const w = (r.warranty || '').toLowerCase().trim();
      return w === 'sì' || w === 'si' || w === 'yes' || w === 's';
    }});
    title = 'Interventi in garanzia';
  }} else if (warrantyValue === 'no') {{
    filtered = reports.filter(r => {{
      const w = (r.warranty || '').toLowerCase().trim();
      return w === 'no' || w === 'n';
    }});
    title = 'Interventi fuori garanzia';
  }} else {{
    filtered = reports.filter(r => {{
      const w = (r.warranty || '').toLowerCase().trim();
      return w !== 'sì' && w !== 'si' && w !== 'yes' && w !== 's' && w !== 'no' && w !== 'n';
    }});
    title = 'Interventi - garanzia non specificata';
  }}
  
  showModalWithReports(title, filtered);
}}

function renderDeviceYearsChart(reports) {{
  const counts = {{}};
  reports.forEach(r => {{
    // Raccogli tutti gli anni dei dispositivi (year1, year2, year3, year4)
    const years = [r.year1, r.year2, r.year3, r.year4].filter(y => y && y !== '');
    // Usa Set per contare ogni anno una sola volta per intervento
    const uniqueYears = [...new Set(years.map(y => String(y).trim()).filter(y => y))];
    uniqueYears.forEach(year => {{
      counts[year] = (counts[year] || 0) + 1;
    }});
  }});
  const data = Object.entries(counts)
    .sort(([a], [b]) => b.localeCompare(a))  // Ordine decrescente per anno
    .map(([k, v]) => ({{ label: k, value: v }}));
  renderBarChart('chartDeviceYears', data, 15, 'chartDeviceYears', true);  // preserveOrder=true per mantenere ordine decrescente
}}

function renderReasonByYearChart(reports) {{
  // Crea matrice: anno -> problema riscontrato -> count
  const matrix = {{}};
  const allProblems = new Set();
  
  reports.forEach(r => {{
    const years = [r.year1, r.year2, r.year3, r.year4].filter(y => y && y !== '');
    const problem = r.problemFound || 'Non specificato';
    allProblems.add(problem);
    
    // Usa Set per contare ogni anno una sola volta per intervento
    const uniqueYears = [...new Set(years.map(y => String(y).trim()).filter(y => y))];
    uniqueYears.forEach(year => {{
      if (!matrix[year]) matrix[year] = {{}};
      matrix[year][problem] = (matrix[year][problem] || 0) + 1;
    }});
  }});
  
  // Genera HTML con barre raggruppate per anno
  const container = document.getElementById('chartReasonByYear');
  const years = Object.keys(matrix).sort().reverse();  // Ordine decrescente
  
  if (years.length === 0) {{
    container.innerHTML = '<div class="no-data">Nessun dato</div>';
    return;
  }}
  
  // Colori per i problemi
  const colors = {{
    'Installazione': '#4CAF50',
    'Guasto meccanico': '#FF5722',
    'Guasto elettrico': '#f44336',
    'Guasto elettronico': '#E91E63',
    'Taratura': '#2196F3',
    'Aggiornamento software': '#00BCD4',
    'Altro': '#FF9800',
    'Non specificato': '#9C27B0'
  }};
  
  let html = '<div style="font-size: 0.75rem;">';
  
  // Legenda
  html += '<div style="display: flex; flex-wrap: wrap; gap: 10px; margin-bottom: 15px;">';
  [...allProblems].sort().forEach(problem => {{
    const color = colors[problem] || '#607D8B';
    html += `<div style="display: flex; align-items: center; gap: 4px;">
      <div style="width: 12px; height: 12px; background: ${{color}}; border-radius: 2px;"></div>
      <span>${{problem}}</span>
    </div>`;
  }});
  html += '</div>';
  
  // Grafico per anno
  years.forEach(year => {{
    const problemCounts = matrix[year];
    const total = Object.values(problemCounts).reduce((a, b) => a + b, 0);
    
    html += `<div style="margin-bottom: 12px;">
      <div style="font-weight: 600; margin-bottom: 4px;">${{year}} <span style="color: #666; font-weight: normal;">(${{total}} interventi)</span></div>
      <div style="display: flex; height: 24px; border-radius: 4px; overflow: hidden;">`;
    
    Object.entries(problemCounts).sort(([a], [b]) => a.localeCompare(b)).forEach(([problem, count]) => {{
      const pct = (count / total * 100).toFixed(1);
      const color = colors[problem] || '#607D8B';
      html += `<div style="width: ${{pct}}%; background: ${{color}}; display: flex; align-items: center; justify-content: center; color: white; font-size: 0.7rem; min-width: 20px;" title="${{problem}}: ${{count}}">${{count}}</div>`;
    }});
    
    html += '</div></div>';
  }});
  
  html += '</div>';
  container.innerHTML = html;
}}

function renderTable(reports) {{
  const tbody = document.querySelector('#reportTable tbody');
  const sorted = [...reports].sort((a, b) => (b.interventionDate || '').localeCompare(a.interventionDate || ''));
  tbody.innerHTML = sorted.map(r => `
    <tr>
      <td>${{formatDate(r.interventionDate)}}</td>
      <td>${{r.companyName || ''}}</td>
      <td>${{r.interventionReason || ''}}</td>
      <td>${{r.hoursWorked || 0}}</td>
      <td>${{r.kilometers || 0}}</td>
      <td>${{(r.grandTotal || 0).toLocaleString('it-IT', {{minimumFractionDigits:2}})}}&euro;</td>
    </tr>
  `).join('');
}}

// --- Helpers ---
function formatDate(iso) {{
  if (!iso) return '';
  const [y, m, d] = iso.split('-');
  return `${{d}}/${{m}}/${{y}}`;
}}
function formatMonth(ym) {{
  const [y, m] = ym.split('-');
  const months = ['Gen','Feb','Mar','Apr','Mag','Giu','Lug','Ago','Set','Ott','Nov','Dic'];
  return `${{months[parseInt(m)-1]}} ${{y}}`;
}}

// --- Init ---
populateFilters();
render();

// Listen for filter changes
document.querySelectorAll('.filters input, .filters select').forEach(el => {{
  el.addEventListener('change', render);
}});
</script>
</body>
</html>"""


def main():
    # Use the folder where this script is located
    folder = Path(__file__).parent

    print(f"📂 Scansione: {folder}")
    print(f"   Cerco file rapporto_*.pdf...")

    reports = load_reports(folder)

    if not reports:
        print("\n⚠️  Nessun PDF con testo trovato.")
        print("   I PDF vecchi (formato immagine) non contengono testo estraibile.")
        print("   Riesporta i rapporti dalla webapp per generare i nuovi PDF testuali.")
        input("\nPremi INVIO per chiudere...")
        sys.exit(0)

    print(f"✅ Estratti dati da {len(reports)} rapporti")

    # Generate dashboard
    html = generate_dashboard_html(reports)
    output_path = folder / "dashboard.html"
    output_path.write_text(html, encoding="utf-8")

    print(f"📊 Dashboard generata: {output_path}")
    print(f"   Aprila con il browser per visualizzare le statistiche.")

    # Auto-open in browser
    try:
        os.startfile(str(output_path))
    except Exception:
        pass

    input("\nPremi INVIO per chiudere...")


if __name__ == "__main__":
    main()
