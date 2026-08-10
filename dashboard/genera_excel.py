"""
Genera un file Excel con i dati estratti dai PDF.
L'Excel può essere modificato manualmente per aggiungere info mancanti.
La dashboard poi legge dall'Excel.

Uso:
    python genera_excel.py
"""

import json
import os
import re
import sys
from pathlib import Path
from datetime import datetime

try:
    import fitz  # PyMuPDF
except ImportError:
    print("❌ PyMuPDF non installato. Esegui: pip install pymupdf")
    input("Premi INVIO per chiudere...")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl non installato. Esegui: pip install openpyxl")
    input("Premi INVIO per chiudere...")
    sys.exit(1)


def normalize_operator(value: str) -> str:
    """Normalizza nomi operatori."""
    if not value:
        return ""
    v = value.strip()
    v_lower = v.lower()
    if 'angelo' in v_lower and 'bocchino' in v_lower:
        return "Angelo Bocchino"
    if 'giacomo' in v_lower and 'mantovani' in v_lower:
        return "Giacomo Mantovani"
    return v.title()


def normalize_company_name(value: str) -> str:
    """Rimuove prefissi comuni dai nomi aziende."""
    if not value:
        return ""
    v = value.strip()
    prefixes = [
        r"^az\.?\s*agricola\s*",
        r"^azienda\s+agricola\s*",
        r"^società\s+agricola\s*",
        r"^soc\.?\s*agricola\s*",
        r"^azienda\s+agricola\s+zootecnica\s*",
        r"^az\.?\s*agr\.?\s*",
    ]
    for pattern in prefixes:
        v = re.sub(pattern, "", v, flags=re.IGNORECASE)
    v = re.sub(r"\s+", " ", v).strip()
    return v.title() if v else value.strip().title()


def extract_report_data(pdf_path: Path) -> dict | None:
    """Estrae i dati da un PDF."""
    try:
        doc = fitz.open(str(pdf_path))
        text = ""
        for page in doc:
            text += page.get_text()
        doc.close()
    except Exception:
        return None

    if not text.strip():
        return None

    data: dict = {"_source": pdf_path.name}

    def find_field(label: str) -> str | None:
        pattern = rf"{re.escape(label)}:\n(.+?)(?:\n|$)"
        match = re.search(pattern, text)
        if not match:
            return None
        value = match.group(1).strip()
        if not value or value.endswith(":") or value in ("Dati Cliente", "Dettagli Intervento", "Dispositivi", "Costi", "Allegati", "Firma e Timbro"):
            return None
        return value

    data["companyName"] = normalize_company_name(find_field("Ragione Sociale") or "")
    data["address"] = find_field("Indirizzo") or find_field("Luogo") or find_field("Luogo Intervento") or ""
    data["phone"] = find_field("Telefono") or ""

    # Date
    date_str = find_field("Data") or find_field("Data Intervento") or ""
    if date_str:
        match = re.match(r"(\d{2})/(\d{2})/(\d{4})", date_str)
        if match:
            data["interventionDate"] = f"{match.group(3)}-{match.group(2)}-{match.group(1)}"
            data["interventionDateDisplay"] = date_str
        else:
            data["interventionDate"] = date_str
            data["interventionDateDisplay"] = date_str
    else:
        data["interventionDate"] = ""
        data["interventionDateDisplay"] = ""

    data["operator1"] = normalize_operator(find_field("Operatore 1") or "")
    data["operator2"] = normalize_operator(find_field("Operatore 2") or "")
    data["interventionLocation"] = find_field("Luogo") or find_field("Luogo Intervento") or ""
    data["requestedBy"] = find_field("Richiesto da") or ""
    
    on_behalf = find_field("Per conto di") or "Fyeld"
    data["onBehalfOf"] = "Fyeld" if re.match(r"(?i)fyeld", on_behalf.strip()) else on_behalf.strip()
    
    data["interventionReason"] = find_field("Motivo") or find_field("Motivo richiesta intervento") or find_field("Motivo Intervento") or ""
    data["problemFound"] = find_field("Problema riscontrato") or ""
    data["description"] = find_field("Descrizione") or find_field("Descrizione dettagliata") or ""
    data["notes"] = find_field("Note") or ""

    # Dispositivi - estrai fino a 4 dispositivi con seriale e anno
    devices = [{"serial": "", "year": ""} for _ in range(4)]
    
    # Metodo 1: campi strutturati "N. Serie:" e "Anno:"
    serial_matches_structured = re.findall(r'N\. Serie:\n(.+?)(?:\n|$)', text)
    year_matches_structured = re.findall(r'Anno:\n(.+?)(?:\n|$)', text)
    
    for i, serial in enumerate(serial_matches_structured[:4]):
        devices[i]["serial"] = serial.strip()
    for i, year in enumerate(year_matches_structured[:4]):
        devices[i]["year"] = year.strip()
    
    # Metodo 2: se non trovati, cerca pattern seriale nel testo libero
    # Pattern: 1ZZ###XX (lettere/numeri) opzionalmente seguito da /YY (anno)
    if not devices[0]["serial"]:
        # Pattern per seriali tipo "1ZZ634SN/17" o "1ZZ771QZ/18" o "1ZZ668FR"
        serial_pattern = r'\b([0-9][A-Z]{2}[0-9]{3}[A-Z]{2})(?:/(\d{2}))?\b'
        matches = re.findall(serial_pattern, text, re.IGNORECASE)
        
        for i, match in enumerate(matches[:4]):
            serial = match[0].upper()
            year_suffix = match[1] if len(match) > 1 else ""
            
            devices[i]["serial"] = serial
            if year_suffix:
                # Converti /17 -> 2017, /18 -> 2018, etc.
                year_num = int(year_suffix)
                if year_num < 50:
                    devices[i]["year"] = f"20{year_suffix}"
                else:
                    devices[i]["year"] = f"19{year_suffix}"
    
    data["serial1"] = devices[0]["serial"]
    data["year1"] = devices[0]["year"]
    data["serial2"] = devices[1]["serial"]
    data["year2"] = devices[1]["year"]
    data["serial3"] = devices[2]["serial"]
    data["year3"] = devices[2]["year"]
    data["serial4"] = devices[3]["serial"]
    data["year4"] = devices[3]["year"]
    
    # Tipo dispositivo (primo)
    device_match = re.search(r'Modello:\n(.+?)(?:\n|$)', text)
    data["deviceType"] = device_match.group(1).strip() if device_match else ""

    # Costi
    hours_match = re.search(r"Ore lavorate\s*(\d+[.,]?\d*)\s*ore", text)
    data["hoursWorked"] = float(hours_match.group(1).replace(",", ".")) if hours_match else 0

    km_match = re.search(r"Chilometri\s*(\d+[.,]?\d*)\s*km", text)
    data["kilometers"] = float(km_match.group(1).replace(",", ".")) if km_match else 0

    # Grand total - supporta vari formati
    total_match = re.search(r"Totale Intervento\s*([\d.,]+)\s*(?:€|EUR)", text)
    if not total_match:
        total_match = re.search(r"Totale Intervento\n([\d.,]+)\s*(?:€|EUR|·)", text)
    
    if total_match:
        total_str = total_match.group(1).strip()
        if '.' in total_str and ',' in total_str:
            if total_str.rfind(',') > total_str.rfind('.'):
                total_str = total_str.replace('.', '').replace(',', '.')
            else:
                total_str = total_str.replace(',', '')
        elif ',' in total_str:
            total_str = total_str.replace(',', '.')
        try:
            data["grandTotal"] = float(total_str)
        except ValueError:
            data["grandTotal"] = 0
    else:
        data["grandTotal"] = 0

    data["hasTravelCost"] = "Ore viaggio" in text

    if data["companyName"] or data["interventionDate"]:
        return data
    return None


def generate_excel(reports: list[dict], output_path: Path):
    """Genera file Excel con i dati dei rapporti."""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapporti"
    
    # Intestazioni
    headers = [
        ("File PDF", 40),
        ("Data", 12),
        ("Azienda", 30),
        ("Indirizzo", 35),
        ("Telefono", 15),
        ("Operatore 1", 18),
        ("Operatore 2", 18),
        ("Per conto di", 12),
        ("Motivo", 15),
        ("Problema riscontrato", 25),
        ("Descrizione", 40),
        ("Tipo dispositivo", 15),
        ("Seriale 1", 15),
        ("Anno 1", 8),
        ("Seriale 2", 15),
        ("Anno 2", 8),
        ("Seriale 3", 15),
        ("Anno 3", 8),
        ("Seriale 4", 15),
        ("Anno 4", 8),
        ("Ricambi utilizzati", 35),
        ("Ore lavorate", 12),
        ("Km", 8),
        ("Totale €", 12),
        ("Viaggio incluso", 14),
        ("Note", 30),
    ]
    
    # Stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Scrivi intestazioni
    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Blocca prima riga
    ws.freeze_panes = "A2"
    
    # Scrivi dati
    for row_idx, report in enumerate(reports, 2):
        row_data = [
            report.get("_source", ""),
            report.get("interventionDateDisplay", ""),
            report.get("companyName", ""),
            report.get("address", ""),
            report.get("phone", ""),
            report.get("operator1", ""),
            report.get("operator2", ""),
            report.get("onBehalfOf", ""),
            report.get("interventionReason", ""),
            report.get("problemFound", ""),
            report.get("description", ""),
            report.get("deviceType", ""),
            report.get("serial1", ""),
            report.get("year1", ""),
            report.get("serial2", ""),
            report.get("year2", ""),
            report.get("serial3", ""),
            report.get("year3", ""),
            report.get("serial4", ""),
            report.get("year4", ""),
            report.get("spareParts", ""),  # Ricambi utilizzati
            report.get("hoursWorked", 0),
            report.get("kilometers", 0),
            report.get("grandTotal", 0),
            "Sì" if report.get("hasTravelCost") else "No",
            report.get("notes", ""),
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    
    # Evidenzia righe con dati mancanti
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for row_idx in range(2, len(reports) + 2):
        address_cell = ws.cell(row=row_idx, column=4)  # Indirizzo
        serial1_cell = ws.cell(row=row_idx, column=13)  # Seriale 1
        km_cell = ws.cell(row=row_idx, column=23)  # Km
        
        # Evidenzia se manca indirizzo
        if not address_cell.value:
            address_cell.fill = yellow_fill
        
        # Evidenzia se manca seriale 1
        if not serial1_cell.value:
            serial1_cell.fill = yellow_fill
        
        # Evidenzia se km = 0
        if km_cell.value == 0:
            km_cell.fill = yellow_fill
    
    # Salva
    wb.save(output_path)
    print(f"✅ Excel generato: {output_path}")
    print(f"   {len(reports)} rapporti esportati")
    print(f"   Le celle gialle indicano dati mancanti da completare")


def complete_missing_data(reports: list[dict]) -> list[dict]:
    """Completa dati mancanti (seriali, anni, indirizzo) da altri rapporti dello stesso cliente."""
    
    # Raggruppa per cliente (nome normalizzato)
    by_company = {}
    for report in reports:
        company = report.get("companyName", "").lower().strip()
        if company:
            if company not in by_company:
                by_company[company] = []
            by_company[company].append(report)
    
    # Per ogni gruppo, raccogli tutti i seriali/anni/indirizzi disponibili
    for company, company_reports in by_company.items():
        # Raccogli tutti i dati disponibili dal gruppo
        all_serials = set()
        all_addresses = set()
        serial_year_map = {}  # seriale -> anno
        
        for r in company_reports:
            for i in range(1, 5):
                serial = r.get(f"serial{i}", "")
                year = r.get(f"year{i}", "")
                if serial:
                    all_serials.add(serial)
                    if year and serial not in serial_year_map:
                        serial_year_map[serial] = year
            
            addr = r.get("address", "")
            if addr:
                all_addresses.add(addr)
        
        # Lista ordinata dei seriali unici
        serials_list = sorted(all_serials)
        best_address = max(all_addresses, key=len) if all_addresses else ""
        
        # Completa i rapporti con dati mancanti
        for r in company_reports:
            # Completa indirizzo se manca
            if not r.get("address") and best_address:
                r["address"] = best_address
            
            # Completa seriali se mancano
            current_serials = [r.get(f"serial{i}", "") for i in range(1, 5)]
            if not any(current_serials):
                # Nessun seriale trovato, usa quelli del gruppo
                for i, serial in enumerate(serials_list[:4], 1):
                    r[f"serial{i}"] = serial
                    if serial in serial_year_map:
                        r[f"year{i}"] = serial_year_map[serial]
            else:
                # Ha alcuni seriali, completa gli anni mancanti
                for i in range(1, 5):
                    serial = r.get(f"serial{i}", "")
                    if serial and not r.get(f"year{i}") and serial in serial_year_map:
                        r[f"year{i}"] = serial_year_map[serial]
    
    return reports


def load_existing_excel(excel_path: Path) -> set:
    """Carica i nomi file già presenti nell'Excel."""
    if not excel_path.exists():
        return set()
    
    existing = set()
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # File PDF
                existing.add(row[0])
        
        wb.close()
    except Exception as e:
        print(f"  ⚠️  Errore lettura Excel esistente: {e}")
    
    return existing


def append_to_excel(excel_path: Path, new_reports: list[dict]):
    """Aggiunge nuovi rapporti all'Excel esistente."""
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Trova l'ultima riga
    last_row = ws.max_row
    
    # Stili
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Aggiungi i nuovi rapporti
    for report in new_reports:
        last_row += 1
        
        row_data = [
            report.get("_source", ""),
            report.get("interventionDateDisplay", ""),
            report.get("companyName", ""),
            report.get("address", ""),
            report.get("phone", ""),
            report.get("operator1", ""),
            report.get("operator2", ""),
            report.get("onBehalfOf", ""),
            report.get("interventionReason", ""),
            report.get("problemFound", ""),
            report.get("description", ""),
            report.get("deviceType", ""),
            report.get("serial1", ""),
            report.get("year1", ""),
            report.get("serial2", ""),
            report.get("year2", ""),
            report.get("serial3", ""),
            report.get("year3", ""),
            report.get("serial4", ""),
            report.get("year4", ""),
            report.get("spareParts", ""),
            report.get("hoursWorked", 0),
            report.get("kilometers", 0),
            report.get("grandTotal", 0),
            "Sì" if report.get("hasTravelCost") else "No",
            report.get("notes", ""),
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=last_row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        
        # Evidenzia celle con dati mancanti
        if not row_data[3]:  # Indirizzo
            ws.cell(row=last_row, column=4).fill = yellow_fill
        if not row_data[12]:  # Seriale 1
            ws.cell(row=last_row, column=13).fill = yellow_fill
        if row_data[22] == 0:  # Km
            ws.cell(row=last_row, column=23).fill = yellow_fill
    
    wb.save(excel_path)
    wb.close()


def main():
    folder = Path(__file__).parent
    
    # Se eseguito dalla cartella dashboard, usa la cartella PDF
    if folder.name == "dashboard":
        folder = Path(r'H:\DG_Assistenza\Assistenze\2026_Assistenze_in campo')
    
    print(f"📂 Scansione: {folder}")
    
    output_path = folder / "rapporti.xlsx"
    
    # Carica nomi file già presenti nell'Excel
    existing_files = load_existing_excel(output_path)
    if existing_files:
        print(f"  📊 {len(existing_files)} rapporti già presenti nell'Excel")
    
    # Trova i PDF nuovi
    pdf_files = sorted(folder.glob("rapporto_*.pdf"))
    new_pdfs = [p for p in pdf_files if p.name not in existing_files]
    
    if not new_pdfs:
        print("\n✅ Nessun nuovo PDF da aggiungere.")
        input("Premi INVIO per chiudere...")
        return
    
    print(f"  ➕ Trovati {len(new_pdfs)} nuovi PDF da aggiungere")
    
    # Estrai dati solo dai nuovi PDF
    new_reports = []
    for pdf_path in new_pdfs:
        data = extract_report_data(pdf_path)
        if data:
            new_reports.append(data)
            print(f"      + {pdf_path.name}")
        else:
            print(f"      ⚠️ Ignorato: {pdf_path.name}")
    
    if not new_reports:
        print("\n⚠️  Nessun nuovo rapporto valido trovato.")
        input("Premi INVIO per chiudere...")
        return
    
    # Completa dati mancanti
    print("  🔄 Completamento dati...")
    new_reports = complete_missing_data(new_reports)
    
    # Se Excel non esiste, crealo da zero
    if not output_path.exists():
        print(f"\n  📝 Creazione nuovo Excel con {len(new_reports)} rapporti...")
        generate_excel(new_reports, output_path)
    else:
        # Aggiungi righe all'Excel esistente
        print(f"\n  📝 Aggiunta {len(new_reports)} righe all'Excel esistente...")
        append_to_excel(output_path, new_reports)
        print(f"✅ Excel aggiornato: {output_path}")
        print(f"   {len(new_reports)} nuovi rapporti aggiunti")
        print(f"   Le celle gialle indicano dati mancanti da completare")
    
    # Apri Excel
    try:
        os.startfile(str(output_path))
    except Exception:
        pass
    
    input("\nPremi INVIO per chiudere...")


if __name__ == "__main__":
    main()
